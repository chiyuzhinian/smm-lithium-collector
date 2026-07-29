"""规范日报生成。读取映射配置，匹配SMM数据和外部数据，生成11列规范报表。"""
from __future__ import annotations
import logging, os, re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
import yaml

log = logging.getLogger("smm_collector.business_report")

REPORT_COLUMNS = [
    "公司", "物料属性", "信息类别", "化学类型", "详细内容",
    "当日均价（YYYY-MM-DD）", "单位", "涨跌",
    "下期价格预测", "数据来源", "备注",
]


def load_mapping(config_root: Path) -> dict:
    """加载业务映射配置。"""
    path = config_root / "business_report_mapping.yaml"
    if not path.exists():
        log.warning("业务映射配置不存在: %s", path)
        return {"company_groups": []}
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_external_data(config_root: Path) -> dict:
    """加载外部数据（Benchmark/客户反馈等）。返回 {external_key: price_dict}。"""
    mapping = load_mapping(config_root)
    ext_cfg = mapping.get("external_sources", {})
    if not ext_cfg.get("enabled", False):
        return {}

    data_dir = config_root / ext_cfg.get("data_dir", "data/external")
    data_dir.mkdir(parents=True, exist_ok=True)
    fname = ext_cfg.get("file_name", "external_prices.xlsx")
    fpath = data_dir / fname
    if not fpath.exists():
        log.info("外部数据文件不存在: %s", fpath)
        return {}

    try:
        if fpath.suffix in (".xlsx", ".xls"):
            df = pd.read_excel(fpath)
        else:
            df = pd.read_csv(fpath, encoding="utf-8-sig")
    except Exception as e:
        log.warning("外部数据读取失败: %s", e)
        return {}

    result = {}
    for _, row in df.iterrows():
        key = str(row.get("external_key", "")).strip()
        if not key:
            continue
        result[key] = {
            "price": _to_dec(row.get("price") or row.get("average_price")),
            "unit": str(row.get("unit", "")).strip(),
            "price_date": str(row.get("price_date", "")).strip(),
            "source": str(row.get("source", "")).strip(),
            "remark": str(row.get("remark", "")).strip(),
        }
    return result


def match_smm_product(smm_rows: list[dict], entry: dict) -> dict | None:
    """在SMM数据中匹配一个规范条目，返回匹配到的行。"""
    source_type = entry.get("source_type", "smm")
    if source_type != "smm":
        return None

    product_key = entry.get("product_key", "")
    product_spec = entry.get("product_spec", "")
    ssm_alias = entry.get("ssm_alias", "")
    category = entry.get("category", "")

    # 期望单位
    expected_unit = entry.get("unit", "")
    candidates = []
    for r in smm_rows:
        pn = str(r.get("product_name", ""))
        spec = str(r.get("specification", ""))
        cat = str(r.get("category", ""))
        unit = str(r.get("unit", ""))

        # 匹配规则：category + product_name + (optional) specification + unit
        if category and cat != category:
            continue
        if product_key and product_key not in pn:
            if ssm_alias and ssm_alias not in pn:
                continue
        if product_spec and product_spec not in spec:
            continue
        # 单位检查：防止元/吨匹配到%或元/Wh等不同单位
        if expected_unit and unit and expected_unit != unit:
            continue
        candidates.append(r)

    if not candidates and ssm_alias:
        for r in smm_rows:
            pn = str(r.get("product_name", ""))
            unit = str(r.get("unit", ""))
            if (ssm_alias in pn or product_key in pn):
                if not expected_unit or not unit or expected_unit.split("/")[-1] == unit.split("/")[-1]:
                    candidates.append(r)

    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) > 1:
        # 多候选：选最新日期
        candidates.sort(key=lambda x: str(x.get("price_date", "")), reverse=True)
        return candidates[0]

    return None


def build_report(smm_rows: list[dict], config_root: Path | None = None,
                 target_date: date | None = None) -> tuple[pd.DataFrame, dict]:
    """生成规范日报 DataFrame 和质量统计。

    Returns: (report_df, quality_stats)
    """
    if config_root is None:
        config_root = Path(__file__).resolve().parents[2] / "config"

    mapping = load_mapping(config_root)
    external_data = load_external_data(config_root)
    company_groups = mapping.get("company_groups", [])

    rows = []
    quality = {
        "total_required": 0, "matched": 0, "missing_smm": 0,
        "missing_external": 0, "ambiguous": 0, "lfp_undifferentiated": 0,
        "stale_prices": 0, "missing_details": [],
    }

    for group in company_groups:
        company = group.get("company", "")
        for entry in group.get("rows", []):
            quality["total_required"] += 1
            detail = entry.get("detail", "")
            source_type = entry.get("source_type", "smm")

            row = {
                "公司": company,
                "物料属性": entry.get("material_attribute", ""),
                "信息类别": entry.get("info_category", ""),
                "化学类型": entry.get("chemistry", ""),
                "详细内容": detail,
                "当日均价（YYYY-MM-DD）": None,
                "单位": entry.get("unit", ""),
                "涨跌": None,
                "下期价格预测": None,
                "数据来源": entry.get("source", ""),
                "备注": "",
            }

            if source_type == "smm":
                matched = match_smm_product(smm_rows, entry)
                if matched:
                    quality["matched"] += 1
                    row["当日均价（YYYY-MM-DD）"] = matched.get("average_price")
                    row["单位"] = matched.get("unit") or row["单位"]
                    price_date = matched.get("price_date")
                    if price_date:
                        td_str = str(target_date) if target_date else ""
                        pd_str = str(price_date)[:10]
                        if pd_str != td_str:
                            quality["stale_prices"] += 1
                            row["备注"] = f"最近有效报价日期：{pd_str}"

                    # 计算涨跌
                    prev = _get_previous_price(smm_rows, entry, price_date)
                    curr = matched.get("average_price")
                    if curr and prev and prev != 0:
                        try:
                            pct = (Decimal(str(curr)) - Decimal(str(prev))) / Decimal(str(prev))
                            row["涨跌"] = float(pct)
                        except Exception:
                            pass
                else:
                    quality["missing_smm"] += 1
                    quality["missing_details"].append(f"SMM缺失: {detail}")
                    row["备注"] = "当前SMM数据中未匹配到此产品"

            elif source_type == "external":
                ext_key = entry.get("external_key", "")
                ext = external_data.get(ext_key)
                if ext and ext.get("price") is not None:
                    quality["matched"] += 1
                    row["当日均价（YYYY-MM-DD）"] = ext["price"]
                    row["单位"] = ext.get("unit") or row["单位"]
                    if ext.get("price_date"):
                        row["备注"] = f"数据日期：{ext['price_date']}"
                    if ext.get("remark"):
                        row["备注"] = (row["备注"] + "; " + ext["remark"]).strip("; ")
                else:
                    quality["missing_external"] += 1
                    quality["missing_details"].append(f"外部缺失: {detail}")
                    row["备注"] = "当日未提供外部数据"

            elif source_type == "pending":
                row["备注"] = "数据来源待确认"
                quality["missing_details"].append(f"待确认: {detail}")

            rows.append(row)

    df = pd.DataFrame(rows, columns=REPORT_COLUMNS)
    return df, quality


def _get_previous_price(smm_rows: list[dict], entry: dict, current_date) -> Decimal | None:
    """获取上一个有效价格日的平均价（用于涨跌计算）。"""
    product_key = entry.get("product_key", "")
    ssm_alias = entry.get("ssm_alias", "")
    category = entry.get("category", "")

    prev = []
    for r in smm_rows:
        pn = str(r.get("product_name", ""))
        cat = str(r.get("category", ""))
        pd_str = str(r.get("price_date", ""))
        if category and cat != category:
            continue
        if product_key not in pn and (not ssm_alias or ssm_alias not in pn):
            continue
        if str(current_date)[:10] == pd_str[:10]:
            continue
        prev.append((pd_str, r.get("average_price")))

    if not prev:
        return None
    prev.sort(key=lambda x: x[0], reverse=True)
    return prev[0][1]


def _to_dec(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val).replace(",", ""))
    except Exception:
        return None


def write_report_sheet(writer, df: pd.DataFrame, sheet_name: str = "规范日报"):
    """将规范日报写入 ExcelWriter。"""
    df.to_excel(writer, index=False, sheet_name=sheet_name)

    # 格式
    from openpyxl.utils import get_column_letter
    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 列宽
    widths = {"A": 12, "B": 8, "C": 12, "D": 8, "E": 42, "F": 16, "G": 8, "H": 12, "I": 12, "J": 18, "K": 30}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # H列百分比格式
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=8)  # H = 涨跌
        if cell.value is not None:
            cell.number_format = '0.00%'

    # 标题行蓝色背景
    from openpyxl.styles import Font, PatternFill
    blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, 12):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = blue_fill
        cell.font = white_font
