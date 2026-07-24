from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook
from smm_collector.exporter import export_daily, _excel_sheet_name, _sorted_cats

def test_multi_category_export(tmp_path):
	cats = ["锂金属", "锂矿", "锂化合物", "正极材料", "废旧锂电池"]
	rows = []
	for c in cats:
		rows.append({"source":"SMM","market":"SMM锂电现货","category":c,
		             "product_name":"x","price_date":date(2026,7,22),
		             "collected_at":datetime.now()})
	xlsx, csv = export_daily(rows, {
		"status":"success","success_categories":cats,"failed_categories":[]},
		tmp_path, date(2026,7,22))
	sheet_names = set(load_workbook(xlsx).sheetnames)
	assert "全部数据" in sheet_names
	assert "采集说明" in sheet_names
	for c in cats:
		assert _excel_sheet_name(c) in sheet_names
	assert "category" in csv.read_text(encoding="utf-8-sig").splitlines()[0]

def test_original_three_categories(tmp_path):
	cats = ["锂金属","锂矿","锂化合物"]
	rows = [{"source":"SMM","market":"SMM锂电现货","category":c,
	         "product_name":"x","price_date":date(2026,7,22),
	         "collected_at":datetime.now()} for c in cats]
	xlsx, csv = export_daily(rows, {
		"status":"success","success_categories":list(cats),"failed_categories":[]},
		tmp_path, date(2026,7,22))
	sheet_names = set(load_workbook(xlsx).sheetnames)
	assert sheet_names == {"全部数据","采集说明"} | {_excel_sheet_name(c) for c in cats}

def test_excel_sheet_name_truncation():
	long_name = "这是一个非常长的分类名称用于测试Excel截断功能一二三四五六七八九十"
	assert len(_excel_sheet_name(long_name)) <= 31

def test_excel_sheet_name_sanitization():
	assert _excel_sheet_name("a:b") == "a_b"
	assert _excel_sheet_name("c/d\\e") == "c_d_e"

def test_sorted_cats():
	import pandas as pd
	df = pd.DataFrame({"category":["B","A","B","C","A"]})
	assert _sorted_cats(df["category"]) == ["B","A","C"]
