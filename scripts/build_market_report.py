"""生成市场价格报表(匹配模板格式,含日期范围和涨跌基准日)。"""
import sqlite3, yaml, sys; from decimal import Decimal; from pathlib import Path
from datetime import date as dt_date
from openpyxl import Workbook; from openpyxl.styles import *
from openpyxl.utils import get_column_letter as L

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

with open(ROOT / "config/business_report_mapping.yaml", encoding="utf-8") as f:
    m = yaml.safe_load(f)

con = sqlite3.connect(str(ROOT / "data/database/smm_lithium.db"))
con.row_factory = sqlite3.Row
all_rows = [dict(r) for r in con.execute("SELECT * FROM lithium_spot_prices").fetchall()]
con.close()

def dec(v):
    if v is None: return None
    try: return Decimal(str(v))
    except: return None

dates = sorted(set(str(r["price_date"])[:10] for r in all_rows if r.get("price_date")))
latest = dates[-1]
month_dates = [d for d in dates if d[:7] == latest[:7]]
period_start = month_dates[0]
print(f"Range: {period_start}~{latest} ({len(month_dates)} days)")

pm = {}
for r in all_rows:
    avg = dec(r.get("average_price"))
    if avg is None: continue
    pn = str(r.get("product_name", "")).strip()
    spec = str(r.get("specification", "")).strip()
    pm.setdefault(pn, {})[str(r.get("price_date", ""))[:10]] = avg
    if spec:
        pm.setdefault(f"{pn}|{spec}", {})[str(r.get("price_date", ""))[:10]] = avg

def find(entry):
    pk = entry.get("product_key", ""); alias = entry.get("ssm_alias", "")
    pspec = entry.get("product_spec", "")
    for name in [pk, alias]:
        if not name: continue
        if name in pm:
            dp = pm[name]
            pds = sorted(dp.keys())
            pp = [v for d, v in dp.items() if d in month_dates]
            if pp:
                avg = sum(pp) / len(pp)
                # 涨跌基准：最早3个数据日期的均价，防止单日波动
                first3_vals = [dp[d] for d in pds[:3] if d in dp]
                baseline = sum(first3_vals) / len(first3_vals) if first3_vals else None
                baseline_dates = f"{pds[0]}~{pds[min(2,len(pds)-1)]}" if len(pds)>=2 else pds[0]
                return avg, baseline, baseline_dates, pds[-1] if pds else None
    for key, dp in pm.items():
        pn_part = key.split("|")[0]; spec_part = key.split("|")[1] if "|" in key else ""
        matched = False
        if pk: matched = matched or (pk in pn_part) or (pn_part in pk)
        if alias: matched = matched or (alias in pn_part) or (pn_part in alias)
        if not matched: continue
        if pspec and len(pspec) > 10 and pspec not in pn_part and pspec not in spec_part: continue
        pds = sorted(dp.keys())
        pp = [v for d, v in dp.items() if d in month_dates]
        if pp:
            avg = sum(pp) / len(pp)
            first3_vals = [dp[d] for d in pds[:3] if d in dp]
            baseline = sum(first3_vals) / len(first3_vals) if first3_vals else None
            baseline_dates = f"{pds[0]}~{pds[min(2,len(pds)-1)]}" if len(pds)>=2 else pds[0]
            return avg, baseline, baseline_dates, pds[-1] if pds else None
    return None, None, None, None

rows = []; filled = has_chg = 0
for g in m.get("company_groups", []):
    co = g.get("company", "")
    for e in g.get("rows", []):
        st = e.get("source_type", "smm")
        if st in ("smm", "additional"):
            avg, first, first_d, last_d = find(e)
        else:
            avg, first, first_d, last_d = None, None, None, None
        chg = None
        if avg is not None and first is not None and first != 0:
            chg = float((avg - first) / first)
        if avg is not None: filled += 1
        if chg is not None: has_chg += 1
        rem = ""
        if st == "external": rem = "当日未提供外部数据"
        elif st == "pending": rem = "数据来源待确认"
        elif avg is None: rem = "数据库中未匹配"
        # 标注实际数据范围与涨跌基准日期（统一列头，详情写入备注）
        if avg is not None:
            parts = []
            if first_d and last_d and first_d != last_d:
                parts.append(f"数据区间{first_d}~{last_d}")
            elif last_d:
                parts.append(f"数据日期{last_d}")
            if first_d:
                parts.append(f"涨跌基准期{first_d}")
            if parts:
                rem = f"{rem}; {'; '.join(parts)}" if rem else "; ".join(parts)
        rows.append({
            "公司": co, "物料属性": e.get("material_attribute", ""),
            "信息类别": e.get("info_category", ""), "化学类型": e.get("chemistry", ""),
            "详细内容": e.get("detail", ""),
            f"期间均价\n({period_start}~{latest})": float(avg) if avg else None,
            "单位": e.get("unit", ""),
            "涨跌\n(期初→期末)": chg,
            "数据来源": e.get("source", ""), "备注": rem,
        })

print(f"Filled: {filled}/47 | Change: {has_chg}")

out = ROOT / "data/exports/市场价格报表_2026-07.xlsx"
wb = Workbook(); ws = wb.active; ws.title = f"{period_start[5:]}-{latest[5:]}"
hdrs = list(rows[0].keys())
for ci, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cur = ""; ms_v = 2
for ri, rd in enumerate(rows):
    r = ri + 2; co = rd.get("公司", "")
    if co and co != cur:
        if cur and r - ms_v > 1: ws.merge_cells(start_row=ms_v, start_column=1, end_row=r - 1, end_column=1)
        cur = co; ms_v = r
    for ci, h in enumerate(hdrs, 1):
        v = rd.get(h); c = ws.cell(row=r, column=ci, value=v if v is not None else None)
        if ci == 6 and v is not None: c.number_format = "#,##0"
        if ci == 8 and v is not None: c.number_format = "0.00%"
        c.font = Font(name="微软雅黑", size=9); c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                          top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
if cur and len(rows) + 1 - ms_v > 1:
    ws.merge_cells(start_row=ms_v, start_column=1, end_row=len(rows) + 1, end_column=1)

for i, w in enumerate([12, 8, 12, 8, 45, 22, 8, 26, 18, 30], 1):
    ws.column_dimensions[L(i)].width = w
ws.row_dimensions[1].height = 55; ws.freeze_panes = "A2"
wb.save(out)
print(f"Saved: {out}")
