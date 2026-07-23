#!/usr/bin/env python
"""SMM 锂电现货价格 — 领导汇报报告生成器。

从 SQLite 数据库读取采集数据，生成结构化的产业链全景 Excel 报告。
用法：
    python scripts/generate_report.py                    # 最新一天
    python scripts/generate_report.py --date 2026-07-22  # 指定日期
    python scripts/generate_report.py --days 7            # 最近7天趋势
"""
from __future__ import annotations
import argparse, sqlite3, sys
from datetime import date, timedelta
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

# ── 产业链结构定义 ──────────────────────────────────────────────
SUPPLY_CHAIN = {
    "上游·矿产资源": ["锂矿", "锂金属", "钴矿", "钴金属", "磷矿", "镍化合物", "锰化合物", "铁源", "碳素", "电炉钢"],
    "上游·钴化合物": ["钴化合物"],
    "中游·锂化合物": ["锂化合物"],
    "中游·正极材料": ["正极材料"],
    "中游·负极材料": ["人造石墨", "天然石墨", "天然石墨负极", "新型负极", "焦类"],
    "中游·电解液": ["电解液", "溶剂及相关原料", "添加剂"],
    "中游·隔膜": ["隔膜"],
    "中游·集流体": ["铜箔", "铝箔"],
    "中游·辅料": ["PVDF", "其他辅料"],
    "中游·磷化工": ["磷化工"],
    "下游·电芯": ["电芯", "储能电芯"],
    "下游·电池系统": ["电池舱", "PACK"],
    "回收·废旧电池": ["废旧锂电池", "未注液电芯价格", "未注液卷芯价格", "废旧正极片及系数"],
    "回收·黑粉": ["废旧锂电黑粉系数指数", "废旧锂电黑粉系数", "废旧锂电黑粉价格"],
    "回收·梯次利用": ["梯次回收价格", "SMM-五矿锂汇通废旧锂电池"],
}

# 核心关注品种（汇报时重点展示）
KEY_PRODUCTS = [
    "电池级碳酸锂", "工业级碳酸锂", "SMM电池级碳酸锂指数",
    "电池级氢氧化锂（粗颗粒）", "电池级氢氧化锂（微粉）",
    "磷酸铁锂（动力型）", "磷酸铁锂（储能型）",
    "三元材料523", "三元材料811",
    "人造石墨（负极）", "天然石墨（负极）",
    "六氟磷酸锂", "电解液（三元）", "电解液（磷酸铁锂）",
    "隔膜（湿法）", "隔膜（干法）",
    "铜箔（8μm）", "铝箔（12μm）",
    "方形电芯（LFP）", "圆柱电芯（18650）",
]

# ── 样式定义 ──────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
DATA_FONT = Font(name="微软雅黑", size=10)
NUMBER_FONT = Font(name="Consolas", size=10)
RED_FONT = Font(name="微软雅黑", size=10, color="CC0000")
GREEN_FONT = Font(name="微软雅黑", size=10, color="007A33")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_style(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_data_style(ws, start_row, end_row, col_count, number_cols=None):
    number_cols = number_cols or set()
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NUMBER_FONT if c in number_cols else DATA_FONT
            cell.alignment = CENTER if c in number_cols else LEFT
            cell.border = THIN_BORDER
            if c in number_cols and isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.font = RED_FONT
                elif cell.value > 0:
                    cell.font = GREEN_FONT


def _auto_width(ws, col_count, max_width=55):
    for c in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, max_width)


def _freeze(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── 数据读取 ──────────────────────────────────────────────────
def load_data(db_path: Path, target_date: date | None = None, days: int = 1):
    """从数据库读取数据。target_date 为 None 时取最新日期。"""
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row

    if target_date:
        date_str = str(target_date)
    else:
        row = con.execute("SELECT MAX(price_date) FROM lithium_spot_prices").fetchone()
        date_str = row[0] if row and row[0] else str(date.today())

    if days > 1:
        end_date = date.fromisoformat(date_str)
        start_date = end_date - timedelta(days=days - 1)
        rows = con.execute(
            "SELECT * FROM lithium_spot_prices WHERE price_date BETWEEN ? AND ? ORDER BY category, product_name",
            (str(start_date), str(end_date))).fetchall()
    else:
        rows = con.execute(
            "SELECT * FROM lithium_spot_prices WHERE price_date = ? ORDER BY category, product_name",
            (date_str,)).fetchall()

    # 读取采集运行记录
    runs = con.execute(
        "SELECT * FROM collection_runs WHERE target_date = ? ORDER BY started_at DESC",
        (date_str,)).fetchall()

    con.close()
    return [dict(r) for r in rows], [dict(r) for r in runs], date_str


def classify_categories(rows: list[dict]) -> dict:
    """按产业链环节分类数据行。"""
    classified = defaultdict(list)
    unclassified = []
    cat_to_section = {}
    for section, cats in SUPPLY_CHAIN.items():
        for cat in cats:
            cat_to_section[cat] = section

    for r in rows:
        section = cat_to_section.get(r["category"])
        if section:
            classified[section].append(r)
        else:
            unclassified.append(r)

    if unclassified:
        classified["其他"] = unclassified
    return dict(classified)


# ── 报告生成 ──────────────────────────────────────────────────
def generate_report(db_path: Path, output_path: Path, target_date=None, days=1):
    rows, runs, date_str = load_data(db_path, target_date, days)
    if not rows:
        print(f"[ERROR] No data for {date_str} in database. Please run collection first.")
        return None

    classified = classify_categories(rows)
    df_all = pd.DataFrame(rows)
    if "collected_at" in df_all.columns:
        df_all["collected_at"] = pd.to_datetime(df_all["collected_at"])

    # 数值列
    num_cols = []
    for c in ["min_price", "max_price", "average_price", "change_value"]:
        if c in df_all.columns:
            df_all[c] = pd.to_numeric(df_all[c], errors="coerce")
            num_cols.append(c)

    print(f"[OK] {len(rows)} records, {df_all['category'].nunique()} categories")
    print(f"[OK] Data date: {date_str}")

    # ── 写入 Excel ──
    tmp = output_path.with_suffix(".tmp.xlsx")
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        # === Sheet 1: 产业链全景 ===
        _write_overview(writer, classified, df_all, date_str, runs, days)

        # === Sheet 2: 核心品种价格 ===
        _write_key_products(writer, df_all, date_str)

        # === Sheet 3+: 各环节明细 ===
        for section in SUPPLY_CHAIN:
            if section in classified:
                safe_name = section.replace("·", "_")[:31]
                _write_section_detail(writer, classified[section], section, safe_name)

        # === 最后: 全部数据 ===
        if len(df_all.columns) > 0:
            export_cols = [c for c in ["category", "product_name", "specification",
                "min_price", "max_price", "average_price", "change_value", "unit",
                "price_date", "validation_status"] if c in df_all.columns]
            df_all[export_cols].to_excel(writer, index=False, sheet_name="全部原始数据")

    # ── 格式化 ──
    _format_workbook(tmp)

    # ── 最终输出 ──
    os_replace = __import__("os").replace
    os_replace(tmp, output_path)
    print(f"[OK] Report saved: {output_path}")
    return output_path


def _write_overview(writer, classified, df_all, date_str, runs, days):
    """产业链全景 Sheet。"""
    meta = {
        "数据来源": "SMM 上海有色网",
        "报告日期": date_str,
        "数据日期范围": f"{date_str}（单日）" if days == 1 else f"{date_str} 前 {days} 天",
        "采集分类数": df_all["category"].nunique(),
        "总记录数": len(df_all),
        "数据完整度": "✅ 完整" if runs and runs[0].get("status") == "success" else "⚠️ 部分",
    }
    if runs:
        meta["最后采集时间"] = runs[0].get("started_at", "")[:19]

    overview_rows = []
    for section in SUPPLY_CHAIN:
        cats = SUPPLY_CHAIN[section]
        section_data = classified.get(section, [])
        if not section_data:
            overview_rows.append({"产业链环节": section, "分类数": len(cats), "数据行数": 0,
                                  "覆盖产品数": 0, "状态": "—"})
            continue
        sec_df = pd.DataFrame(section_data)
        products = sec_df["product_name"].nunique() if "product_name" in sec_df.columns else 0
        overview_rows.append({
            "产业链环节": section, "分类数": len(cats), "数据行数": len(section_data),
            "覆盖产品数": products, "状态": "✅" if len(section_data) > 0 else "—",
        })

    df_ov = pd.DataFrame(overview_rows)
    df_ov.to_excel(writer, index=False, sheet_name="产业链全景", startrow=6)

    # 元数据放在上方
    ws = writer.sheets["产业链全景"]
    ws.merge_cells("A1:E1"); ws["A1"] = "SMM 锂电现货价格 — 产业链全景报告"; ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:E2"); ws["A2"] = f"报告日期：{date_str}  |  数据来源：SMM 上海有色网  |  采集分类数：{df_all['category'].nunique()}"; ws["A2"].font = DATA_FONT
    row = 4
    for k, v in meta.items():
        ws.cell(row=row, column=1, value=k).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(v)).font = DATA_FONT
        row += 1
    _apply_header_style(ws, 7, 5)
    _auto_width(ws, 5)
    _freeze(ws)


def _write_key_products(writer, df_all, date_str):
    """核心品种价格 Sheet — 只展示最重要的产品。"""
    key_rows = df_all[df_all["product_name"].isin(KEY_PRODUCTS)].copy()
    if key_rows.empty:
        # 若没有精确匹配，展示每个分类的第一条
        key_rows = df_all.groupby("category").first().reset_index()

    display_cols = [c for c in ["category", "product_name", "specification",
        "min_price", "max_price", "average_price", "change_value", "unit"]
        if c in key_rows.columns]
    key_rows = key_rows[display_cols].sort_values(["category", "product_name"])

    key_rows.to_excel(writer, index=False, sheet_name="核心品种价格", startrow=3)
    ws = writer.sheets["核心品种价格"]
    ws.merge_cells("A1:H1"); ws["A1"] = f"SMM 锂电核心品种现货价格 — {date_str}"; ws["A1"].font = TITLE_FONT
    _apply_header_style(ws, 4, len(display_cols))
    _auto_width(ws, len(display_cols))
    _freeze(ws)


def _write_section_detail(writer, rows, section_name, sheet_name):
    """某一产业链环节的详细数据 Sheet。"""
    df = pd.DataFrame(rows)
    display_cols = [c for c in ["category", "product_name", "specification",
        "min_price", "max_price", "average_price", "change_value", "unit",
        "price_date", "validation_status"] if c in df.columns]
    df = df[display_cols].sort_values(["category", "product_name"])

    df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
    ws = writer.sheets[sheet_name]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_cols))
    ws["A1"] = f"📋 {section_name}（{len(rows)} 条记录）"
    ws["A1"].font = SECTION_FONT
    _apply_header_style(ws, 3, len(display_cols))
    _auto_width(ws, len(display_cols))
    _freeze(ws)


def _format_workbook(path):
    """对 openpyxl 工作簿进行全局格式化。"""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        # 冻结首行
        if ws.freeze_panes is None:
            ws.freeze_panes = "A2"
    wb.save(path)


# ── CLI ───────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="SMM 锂电现货领导汇报报告生成器")
    p.add_argument("--date", type=date.fromisoformat, help="报告日期（默认最新）")
    p.add_argument("--days", type=int, default=1, help="覆盖天数（默认1天）")
    p.add_argument("--output", type=Path, help="输出路径（默认 data/exports/领导汇报/）")
    p.add_argument("--db", type=Path, default=ROOT / "data/database/smm_lithium.db",
                   help="数据库路径")
    args = p.parse_args()

    if not args.db.exists():
        print(f"[ERROR] Database not found: {args.db}")
        sys.exit(1)

    if args.output:
        out = args.output
        out.parent.mkdir(parents=True, exist_ok=True)
    else:
        today = date.today()
        out_dir = ROOT / "data/exports" / "summary"
        out_dir.mkdir(parents=True, exist_ok=True)
        date_label = str(args.date or "latest")
        days_label = f"_{args.days}days" if args.days > 1 else ""
        out = out_dir / f"SMM_lithium_supply_chain_report_{date_label}{days_label}.xlsx"

    result = generate_report(args.db, out, args.date, args.days)
    if result is None:
        sys.exit(1)


if __name__ == "__main__":
    main()
